import os
import csv
import re
from datetime import datetime
from openpyxl import load_workbook

src_dir = 'src_files'
dst_dir = 'csv_files'
os.makedirs(dst_dir, exist_ok=True)

SOURCE_ID = '5'

# 正则模式
phone_pattern = re.compile(r'^1\d{10}$')
email_pattern = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')
id_pattern = re.compile(r'^\d{17}[\dXx]$')
punctuation_pattern = re.compile(r'[^\w\u4e00-\u9fa5]')  # 非中文、非字母数字字符

# 用于身份证校验出生日期
def validate_id(id_number):
    if not id_number or not id_pattern.match(id_number):
        return None
    birth = id_number[6:14]
    try:
        datetime.strptime(birth, "%Y%m%d")
    except ValueError:
        return None
    return id_number

# 清洗姓名
def clean_name(name):
    if not name:
        return None
    name = str(name).strip()
    name = re.sub(r'[\d%s]' % re.escape('!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'), '', name)
    return name if name else None

# 清洗手机号
def clean_phone(phone):
    if not phone:
        return None
    phone = str(phone).replace(' ', '')
    return phone if phone_pattern.match(phone) else None

# 清洗邮箱
def clean_email(email):
    if not email:
        return None
    email = str(email).strip()
    return email if email_pattern.match(email) else None

# 清洗地址
def clean_address(address):
    if not address:
        return None
    address = str(address).strip()
    # 去掉标点符号和空格
    address = re.sub(r'[^\w\u4e00-\u9fa5]', '', address)
    # 删除逗号
    address = address.replace(',', '')
    # 如果开头是数字或字母，后面跟汉字，去掉开头
    if re.match(r'^[A-Za-z0-9][\u4e00-\u9fa5]', address):
        address = address[1:]
    return address if address else None

# 清洗配置
def clean_vehicle(vehicle):
    if not vehicle:
        return None
    return str(vehicle).replace(',', '').strip() or None

for filename in os.listdir(src_dir):
    if not filename.endswith('.xlsx'):
        continue

    src_path = os.path.join(src_dir, filename)
    dst_path = os.path.join(dst_dir, filename.rsplit('.', 1)[0] + '.csv')

    wb = load_workbook(src_path, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    headers_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

    valid_count = 0
    skipped_count = 0

    with open(dst_path, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)

        for row in rows_iter:
            # 读取源字段
            name = row[headers_map['姓名']] if '姓名' in headers_map else None
            id_card = row[headers_map['身份证']] if '身份证' in headers_map else None
            phone = row[headers_map['手机']] if '手机' in headers_map else None
            email = row[headers_map['邮箱']] if '邮箱' in headers_map else None
            address = row[headers_map['地址']] if '地址' in headers_map else None
            vehicle = row[headers_map['配置']] if '配置' in headers_map else None

            # 清洗
            name = clean_name(name)
            id_card = validate_id(id_card)
            phone = clean_phone(phone)
            email = clean_email(email)
            address = clean_address(address)
            vehicle = clean_vehicle(vehicle)

            # 行过滤
            if not any([id_card, phone, email]):
                skipped_count += 1
                continue

            dst_row = [
                id_card,   # 身份证号
                name,      # 姓名
                None,      # 收件人
                None,      # 昵称
                phone,     # 手机号
                address,   # 地址
                vehicle,   # 车辆
                email,     # 邮箱
                None,      # QQ
                None,      # weibo
                None,      # 联系人
                None,      # 企业名称
                SOURCE_ID  # 来源ID
            ]
            writer.writerow(dst_row)
            valid_count += 1

    print(f"[INFO] 文件: {filename} 处理完成")
    print(f"[INFO] 有效行: {valid_count}, 跳过行: {skipped_count}")
    print("-" * 50)

print("全部文件处理完成")