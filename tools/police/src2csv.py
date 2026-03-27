import os
import csv
import re
from datetime import datetime
from openpyxl import load_workbook

# 目录
src_dir = 'src_files'
dst_dir = 'csv_files'
os.makedirs(dst_dir, exist_ok=True)

SOURCE_ID = '7'

# 正则
phone_pattern = re.compile(r'^1\d{10}$')

# ========================
# 基础清洗（所有字段先走这里）
# ========================
def clean_base(val):
    if val is None:
        return None
    val = str(val)
    val = val.replace('\t', '').replace('\n', '').replace('\r', '').replace('\xa0', '')
    return val.strip()

# ========================
# 身份证
# ========================
def validate_id(id_number):
    id_number = clean_base(id_number)
    if not id_number:
        return None

    id_number = re.sub(r'[^0-9Xx]', '', id_number)

    if not re.fullmatch(r'\d{17}[\dXx]', id_number):
        return None

    birth = id_number[6:14]
    try:
        datetime.strptime(birth, "%Y%m%d")
    except ValueError:
        return None

    return id_number

# ========================
# 姓名
# ========================
def clean_name(name):
    name = clean_base(name)
    if not name:
        return None

    name = name.replace(' ', '')
    name = re.sub(r'[\d!"#$%&\'()*+,-./:;<=>?@[\\\]^_`{|}~]', '', name)

    return name if name else None

# ========================
# 手机号
# ========================
def clean_phone(phone):
    phone = clean_base(phone)
    if not phone:
        return None

    phone = re.sub(r'\D', '', phone)
    return phone if phone_pattern.match(phone) else None

# ========================
# 地址
# ========================
def clean_address(address):
    address = clean_base(address)
    if not address:
        return None

    address = address.replace(',', '').replace('·', '')

    if re.match(r'^[A-Za-z0-9][\u4e00-\u9fa5]', address):
        address = address[1:]

    return address if address else None

# ========================
# 主处理
# ========================
for filename in os.listdir(src_dir):
    if not filename.endswith('.xlsx'):
        continue

    src_path = os.path.join(src_dir, filename)
    dst_path = os.path.join(dst_dir, filename.rsplit('.', 1)[0] + '.csv')

    wb = load_workbook(src_path, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # 表头
    headers = next(rows_iter)
    headers_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

    # 统计字段
    required_fields = ['名字', '身份证', '手机号', '地址']
    found_fields = [f for f in required_fields if f in headers_map]

    print(f"[INFO] 文件: {filename}")
    print(f"[INFO] 找到字段数量: {len(found_fields)} -> {found_fields}")

    valid_count = 0
    skipped_count = 0

    with open(dst_path, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)

        for row in rows_iter:
            name = row[headers_map['名字']] if '名字' in headers_map else None
            id_card = row[headers_map['身份证']] if '身份证' in headers_map else None
            phone = row[headers_map['手机号']] if '手机号' in headers_map else None
            address = row[headers_map['地址']] if '地址' in headers_map else None

            # 清洗
            name = clean_name(name)
            id_card = validate_id(id_card)
            phone = clean_phone(phone)
            address = clean_address(address)

            # 过滤
            if not any([id_card, phone]):
                skipped_count += 1
                continue

            dst_row = [
                id_card,
                name,
                None,
                None,
                phone,
                address,
                None,
                None,
                None,
                None,
                None,
                None,
                SOURCE_ID
            ]

            writer.writerow(dst_row)
            valid_count += 1

    print(f"[INFO] 输出有效行数: {valid_count}")
    print(f"[INFO] 跳过行数: {skipped_count}")
    print("-" * 50)

print("全部文件处理完成")