import os
import csv
import re
from openpyxl import load_workbook

src_dir = 'src_files'
dst_dir = 'csv_files'
os.makedirs(dst_dir, exist_ok=True)

# 手机号校验
phone_pattern = re.compile(r'^1\d{10}$')

# 匹配 “4月2日” 这种
date_pattern = re.compile(r'^\d+月\d+日$')

def clean_text(val):
    if val is None:
        return None
    val = str(val).replace(' ', '').strip()
    return val if val else None

def is_all_digit(val):
    return val.isdigit() if val else False

for filename in os.listdir(src_dir):
    if not filename.endswith('.xlsx'):
        continue

    src_path = os.path.join(src_dir, filename)
    dst_path = os.path.join(dst_dir, filename.rsplit('.', 1)[0] + '.csv')

    wb = load_workbook(src_path, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)

    # 读取表头
    headers = next(rows)
    headers_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

    required_fields = ['联系人', '联系方式', '企业名称', '企业地址']
    found_fields = [f for f in required_fields if f in headers_map]
    missing_fields = [f for f in required_fields if f not in headers_map]

    print(f"[INFO] 文件: {filename}")
    print(f"[INFO] 找到字段: {len(found_fields)}/4 -> {found_fields}")
    if missing_fields:
        print(f"[WARN] 缺失字段: {missing_fields}")

    valid_count = 0

    with open(dst_path, 'w', newline='', encoding='utf-8') as f_out:
        writer = csv.writer(f_out)

        for row in rows:
            # 取值（可能不存在）
            contact = row[headers_map['联系人']] if '联系人' in headers_map else None
            phone = row[headers_map['联系方式']] if '联系方式' in headers_map else None
            company = row[headers_map['企业名称']] if '企业名称' in headers_map else None
            address = row[headers_map['企业地址']] if '企业地址' in headers_map else None

            # === 清洗 ===

            # 联系人
            contact = clean_text(contact)
            if contact and is_all_digit(contact):
                contact = None

            # 企业名称
            company = clean_text(company)
            if company and is_all_digit(company):
                company = None

            # 手机号（关键校验）
            phone = clean_text(phone)
            if not phone or not phone_pattern.match(phone):
                continue  # 丢弃整行

            # 地址
            address = clean_text(address)
            if address:
                if is_all_digit(address) or date_pattern.match(address):
                    address = None

            # 构建目标行
            dst_row = [
                None,      # 身份证号
                None,      # 姓名
                None,      # 收件人
                None,      # 昵称
                phone,     # 手机号
                address,   # 地址
                None,      # 车辆
                None,      # 邮箱
                None,      # QQ
                None,      # weibo
                contact,   # 联系人
                company,   # 企业名称
                '4'        # 来源ID
            ]

            writer.writerow(dst_row)
            valid_count += 1

    print(f"[INFO] 文件处理完成: {filename}")
    print(f"[INFO] 合法输出行数: {valid_count}")
    print("-" * 50)

print("全部文件处理完成")